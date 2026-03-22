"""
monte_carlo.py  —  Personal Finance Monte Carlo Simulator
==========================================================
USAGE
-----
Save this file in the SAME folder as your Excel file, then run:

    python monte_carlo.py

The script reads your best allocation from the Optimizer Engine,
simulates N independent 35-year life paths by randomly drawing
year-by-year economic scenarios (exactly as your Assumptions sheet
does with RAND()), and reports the probability of each goal being met.

REQUIREMENTS
------------
    pip install openpyxl
"""

import random
import os
from openpyxl import load_workbook

# ─── CONFIG ──────────────────────────────────────────────────────────────────
N_SIMS       = 1000      # number of Monte Carlo runs (increase for more precision)
EXCEL_FILE   = "PersonalFinanceWithMacro.xlsm"   # must be in same folder as this script
# ─────────────────────────────────────────────────────────────────────────────


def get_rates(rand1, rand2):
    """
    Replicate the Assumptions Sheet lookup logic exactly.
    rand1  → picks income/inflation scenario  (Good / Mid / Bad)
    rand2  → picks investment return tier     (Below Avg / Avg / Above Avg)
    Returns a dict of all rates for the year.
    """
    # Income scenario (Assumptions rows 14-16, col probabilities 0.8/0.1/0.1)
    if rand1 < 0.8:                      # Good
        salary_g   = 0.10
        infl_gen   = 0.05
        infl_food  = 0.04
        infl_ret   = 0.08
        infl_med   = 0.04
        infl_house = 0.05
        adhoc_exp  = 10000
    elif rand1 < 0.9:                    # Mid
        salary_g   = 0.07
        infl_gen   = 0.06
        infl_food  = 0.05
        infl_ret   = 0.05
        infl_med   = 0.04
        infl_house = 0.04
        adhoc_exp  = 15000
    else:                                # Bad
        salary_g   = 0.04
        infl_gen   = 0.08
        infl_food  = 0.07
        infl_ret   = 0.10
        infl_med   = 0.04
        infl_house = 0.08
        adhoc_exp  = 30000

    # Return tier (Assumptions rows 14-16, return prob cols 0.8/0.1/0.1)
    if rand2 < 0.8:                      # Below Avg
        lc_ret  = 0.14
        mc_ret  = 0.16
        sc_ret  = 0.20
    elif rand2 < 0.9:                    # Avg
        lc_ret  = 0.12
        mc_ret  = 0.14
        sc_ret  = 0.16
    else:                                # Above Avg
        lc_ret  = 0.09
        mc_ret  = 0.12
        sc_ret  = 0.14

    fd_ret  = 0.05
    sav_ret = 0.03

    return {
        "salary_g":   salary_g,
        "infl_gen":   infl_gen,
        "infl_food":  infl_food,
        "infl_ret":   infl_ret,
        "infl_med":   infl_med,
        "infl_house": infl_house,
        "adhoc_exp":  adhoc_exp,
        "lc_ret":     lc_ret,
        "mc_ret":     mc_ret,
        "sc_ret":     sc_ret,
        "fd_ret":     fd_ret,
        "sav_ret":    sav_ret,
    }


def simulate_one_path(lc_pct, mc_pct, sc_pct, goals):
    """
    Simulate one 35-year financial path (age 20 → 54).
    Returns dict of {goal_name: True/False} — whether corpus >= target at goal age.

    Exactly mirrors the Probability Engine's Calc Engine logic:
      - Income grows by salary_g each year
      - Expenses grow by their respective inflation rates
      - Surplus after expenses → invest fixed amount (grows 10%/yr)
      - Investment split: lc_pct / mc_pct / sc_pct
      - Savings account grows at sav_ret; excess swept to FD
      - Fund corpus withdrawn at goal age if goal assigned to that fund
    """

    # ── INITIAL STATE (from Prob Engine row 29 / Details sheet) ──────────────
    salary      = 720_000
    rent_inc    = 0
    interest_inc= 6_000
    biz_inc     = 0

    emi         = 0
    entertainment = 18_000
    food        = 36_000
    recreation  = 12_000
    medical     = 6_000
    luxury      = 24_000
    transport   = 18_000
    household   = 84_000
    insurance   = 12_000
    adhoc       = 10_000   # baseline; overridden by scenario each year

    min_invest  = 40_000   # minimum annual investment
    invest_incr = 0.10     # investment amount grows 10%/yr

    sav_bal     = 80_000   # opening savings account balance
    min_balance = 100_000  # keep at least this in savings
    fd_sweep    = 300_000  # auto-sweep this amount to FD when possible
    fd_rate     = 0.05
    sav_rate    = 0.03

    lc_corpus   = 0.0
    mc_corpus   = 0.0
    sc_corpus   = 0.0
    fd_corpus   = 0.0

    current_invest = min_invest

    # ── GOAL TRACKING ─────────────────────────────────────────────────────────
    # goals = list of dicts: {name, age, lc_goal, mc_goal, sc_goal}
    goal_met = {g["name"]: False for g in goals}

    # ── SIMULATE YEAR BY YEAR (age 20 to 54 = 35 years) ─────────────────────
    for yr in range(35):
        age = 20 + yr

        # Draw random scenario for this year
        rand1 = random.random()
        rand2 = random.random()
        r = get_rates(rand1, rand2)

        # ── INCOME ───────────────────────────────────────────────────────────
        if yr > 0:
            salary       *= (1 + r["salary_g"])
            interest_inc *= 1.05   # interest grows 5% always

        total_income = salary + rent_inc + interest_inc + biz_inc

        # ── EXPENSES (grow with inflation) ───────────────────────────────────
        if yr > 0:
            entertainment *= (1 + r["infl_gen"])
            food          *= (1 + r["infl_food"])
            recreation    *= (1 + r["infl_gen"])
            medical       *= (1 + r["infl_med"])
            luxury        *= (1 + r["infl_ret"])
            transport     *= (1 + r["infl_gen"])
            household     *= (1 + r["infl_house"])
            # insurance stays fixed at 12000 (matches model)

        adhoc = r["adhoc_exp"]   # scenario-driven, not compounded

        total_expenses = (emi + entertainment + food + recreation + medical
                          + luxury + transport + household + insurance + adhoc)

        # ── SURPLUS & INVESTMENT ──────────────────────────────────────────────
        balance = total_income - total_expenses

        if balance >= min_invest:
            invest_amt = current_invest
        else:
            invest_amt = 0   # can't invest if surplus too low

        cashflow_to_savings = balance - invest_amt


        # ── GROW FUNDS: add new investment, apply end-of-year returns ────────
        lc_add = invest_amt * lc_pct
        mc_add = invest_amt * mc_pct
        sc_add = invest_amt * sc_pct

        lc_corpus = (lc_corpus + lc_add) * (1 + r["lc_ret"])
        mc_corpus = (mc_corpus + mc_add) * (1 + r["mc_ret"])
        sc_corpus = (sc_corpus + sc_add) * (1 + r["sc_ret"])

        # ── CHECK GOALS at END of goal year (after full-year growth) ─────────
        # Mirrors Excel: corpus at end-of-year vs target, then withdraw
        for g in goals:
            if g["age"] == age:
                if g["lc_goal"] > 0:
                    if lc_corpus >= g["lc_goal"]:
                        lc_corpus -= g["lc_goal"]
                        goal_met[g["name"]] = True
                elif g["mc_goal"] > 0:
                    if mc_corpus >= g["mc_goal"]:
                        mc_corpus -= g["mc_goal"]
                        goal_met[g["name"]] = True
                elif g["sc_goal"] > 0:
                    if sc_corpus >= g["sc_goal"]:
                        sc_corpus -= g["sc_goal"]
                        goal_met[g["name"]] = True
                else:
                    goal_met[g["name"]] = True   # no ₹ target → always met


        # ── SAVINGS ACCOUNT ───────────────────────────────────────────────────
        sav_bal = sav_bal * (1 + sav_rate) + cashflow_to_savings

        # Sweep excess to FD if savings > threshold
        if sav_bal > fd_sweep + min_balance:
            sweep_amt = fd_sweep
            sav_bal  -= sweep_amt
            fd_corpus = (fd_corpus + sweep_amt) * (1 + fd_rate)
        else:
            fd_corpus = fd_corpus * (1 + fd_rate)

        # ── INVESTMENT INCREMENT FOR NEXT YEAR ────────────────────────────────
        if yr > 0:
            current_invest *= (1 + invest_incr)

    return goal_met


def run_simulation():
    # ── LOAD EXCEL & READ BEST ALLOCATION ─────────────────────────────────────
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, EXCEL_FILE)

    if not os.path.exists(excel_path):
        print(f"\nERROR: Cannot find '{EXCEL_FILE}' in {script_dir}")
        print("Make sure the Excel file and this script are in the same folder.")
        return

    print(f"\nLoading: {EXCEL_FILE}")
    wb = load_workbook(excel_path, keep_vba=True, data_only=True)

    ws_opt = wb["Optimizer Engine"]
    lc_pct = float(ws_opt["H25"].value)
    mc_pct = float(ws_opt["H26"].value)
    sc_pct = float(ws_opt["H27"].value)

    print(f"Best allocation from Optimizer Engine:")
    print(f"  Large Cap : {lc_pct:.0%}")
    print(f"  Mid Cap   : {mc_pct:.0%}")
    print(f"  Small Cap : {sc_pct:.0%}")

    if abs(lc_pct + mc_pct + sc_pct - 1.0) > 0.01:
        print("\nWARNING: Allocations don't sum to 100%. Check H25/H26/H27 in Optimizer Engine.")

    # ── READ GOALS FROM PROBABILITY ENGINE ────────────────────────────────────
    ws_pe = wb["Probablity Engine"]
    goals = []
    goal_rows = {8: "Own Health Fund", 9: "Marriage", 10: "Children Education",
                 11: "House",          12: "Car",      13: "Abroad Vacation"}

    for row_num, name in goal_rows.items():
        at_age  = ws_pe.cell(row_num, 40).value   # col 40 = at age
        lc_goal = ws_pe.cell(row_num, 42).value or 0   # col 42 = LC goal ₹
        mc_goal = ws_pe.cell(row_num, 48).value or 0   # col 48 = MC goal ₹
        sc_goal = ws_pe.cell(row_num, 54).value or 0   # col 54 = SC goal ₹
        goals.append({"name": name, "age": at_age,
                       "lc_goal": lc_goal, "mc_goal": mc_goal, "sc_goal": sc_goal})

    print(f"\nGoals loaded:")
    for g in goals:
        total_target = g["lc_goal"] + g["mc_goal"] + g["sc_goal"]
        fund = ("LC" if g["lc_goal"] > 0 else
                "MC" if g["mc_goal"] > 0 else
                "SC" if g["sc_goal"] > 0 else "None")
        print(f"  {g['name']:<22} age {g['age']:>2}   target ₹{total_target:>12,.0f}  fund={fund}")

    # ── RUN MONTE CARLO ────────────────────────────────────────────────────────
    print(f"\nRunning {N_SIMS:,} simulations ...")

    counters   = {g["name"]: 0 for g in goals}
    all_met_ct = 0

    for sim in range(N_SIMS):
        if (sim + 1) % 200 == 0:
            print(f"  {sim+1:>5} / {N_SIMS} done ...")

        result = simulate_one_path(lc_pct, mc_pct, sc_pct, goals)

        all_met = True
        for g in goals:
            if result[g["name"]]:
                counters[g["name"]] += 1
            else:
                all_met = False

        if all_met:
            all_met_ct += 1

    # ── PRINT RESULTS ─────────────────────────────────────────────────────────
    print("\n" + "=" * 55)
    print("  MONTE CARLO RESULTS")
    print(f"  Allocation: Large {lc_pct:.0%}  |  Mid {mc_pct:.0%}  |  Small {sc_pct:.0%}")
    print(f"  Simulations: {N_SIMS:,}")
    print("=" * 55)
    print(f"  {'Goal':<24}  {'Times Met':>10}  {'Probability':>12}")
    print("  " + "-" * 51)
    for g in goals:
        n    = counters[g["name"]]
        prob = n / N_SIMS
        bar  = "█" * int(prob * 20)
        print(f"  {g['name']:<24}  {n:>10,}  {prob:>11.1%}  {bar}")
    print("  " + "-" * 51)
    n    = all_met_ct
    prob = n / N_SIMS
    bar  = "█" * int(prob * 20)
    print(f"  {'ALL Goals Met':<24}  {n:>10,}  {prob:>11.1%}  {bar}")
    print("=" * 55)

    # ── WRITE RESULTS BACK TO EXCEL ───────────────────────────────────────────
    print(f"\nWriting results to Probability Engine (columns BN:BP) ...")

    # Col BN=66, BO=67, BP=68
    ws_pe.cell(1,  66).value = "=== Monte Carlo Results ==="
    ws_pe.cell(2,  66).value = f"Allocation: Large {lc_pct:.0%} | Mid {mc_pct:.0%} | Small {sc_pct:.0%}"
    ws_pe.cell(3,  66).value = f"Simulations: {N_SIMS:,}"
    ws_pe.cell(7,  66).value = "Goal"
    ws_pe.cell(7,  67).value = f"Times Met (of {N_SIMS})"
    ws_pe.cell(7,  68).value = "Probability"

    for i, g in enumerate(goals):
        row = 8 + i
        ws_pe.cell(row, 66).value = g["name"]
        ws_pe.cell(row, 67).value = counters[g["name"]]
        ws_pe.cell(row, 68).value = round(counters[g["name"]] / N_SIMS, 4)

    ws_pe.cell(14, 66).value = "ALL Goals Met"
    ws_pe.cell(14, 67).value = all_met_ct
    ws_pe.cell(14, 68).value = round(all_met_ct / N_SIMS, 4)

    # Save the file
    wb.save(excel_path)
    print(f"Done. Results saved to '{EXCEL_FILE}'")
    print("Open Probability Engine → columns BN:BP to see the results.\n")


if __name__ == "__main__":
    run_simulation()
